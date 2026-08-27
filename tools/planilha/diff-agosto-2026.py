# -*- coding: utf-8 -*-
"""Diferença linha a linha entre a conciliação 34 do sistema e a aba
Agosto-2026 da planilha do Itaú (Acop Files).

O casamento é por (dia, sinal, valor). O que não casa no mesmo dia é procurado
numa janela de +/- 5 dias, que é como se distingue "dinheiro faltando" de
"data diferente".
"""
import io
import sys
from collections import Counter
from datetime import date

import openpyxl

CAMINHO = r'K:\GERAL\TI\Conciliação Itaú Acop Files.xlsx'
ABA = 'Agosto-2026'

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Colado da tela da conciliacao 34 (dia + sinal + valor)
SISTEMA = """
03-124.00 03+2823.19 03+169.86 03+101.46 03+3673.08 03+736.44 03+63822.91
03+9175.70 03+8642.78 03+2228.70 03+534.66 03+13230.48 03+220.47 03+803.70
03+159558.07 03+3182.63 03+699.36 03+288.94 03+1080.00 03+5095.78 03+3142.45
03+740.00 03+180.00 04-1151.50 04-2400.92 04-15587.48 04-4113.30 04+685.14
04+229.53 04+978.97 04+238.26 04+6664.56 04+1900.58 04+9116.64 04+1742.50
04+956.31 04+765.99 04+36566.09 04+4358.78 04+2838.59 04+610.84 04+67.26
04+393.30 04+253.08 04-190000.00 04-15.92 04-151.24 04+0.64 04-4395.00
04-1781.00 04-1166.00 04-1318.00 04-1166.00 04-1662.00 04-1166.00 05-19942.64
05-109.09 05-50000.00 05+0.03 06-134.57 06-24000.06 07-80000.00 07+0.37
10-262.02 10-5953.34 10-55.60 10-6128.46 10-1485.00 10-2465.90 10+510.00
10-50000.00 10+0.63 11-852.39 11+0.02 12-577.00 12+0.02 14-50000.00 17+228.80
17+45531.60 17+879.53 17-11014.51 17-179.00 17-509.60 17-13167.22 17+0.49
18+3170.70 18-1593.07 18-1087.69 18-399.00 18-2917.04 18-34719.62 18-210.00
18+1.08 19+587.18 19-14470.88 19-1933.80 19+0.07 19-6040.20 19-1793.80
19-1193.80 19-1193.80 19-1193.80 19-1193.80 19-1193.80 19-1295.80 20+39635.53
20+8327.91 20-163.60 20-15000.00 20+0.11 24-5477.75 24-1010.63 24-3000.00
24-20000.00 24+0.07
"""


def cents(v):
    if v is None:
        return 0
    if isinstance(v, str):
        v = v.strip().replace('.', '').replace(',', '.')
        return int(round(float(v) * 100)) if v else 0
    return int(round(float(v) * 100))


# --- lado sistema ---
sistema = []
for tok in SISTEMA.split():
    dia = int(tok[:2])
    sinal = tok[2]
    valor = int(round(float(tok[3:]) * 100))
    sistema.append((dia, valor if sinal == '+' else -valor))

# --- lado planilha ---
wb = openpyxl.load_workbook(CAMINHO, data_only=True)
ws = wb[ABA]

planilha = []
rotulo = {}
for r in range(5, ws.max_row + 1):
    d = ws.cell(row=r, column=1).value
    if d is None:
        continue
    e, s = cents(ws.cell(row=r, column=5).value), cents(ws.cell(row=r, column=6).value)
    if e == 0 and s == 0:
        continue
    hist = str(ws.cell(row=r, column=4).value or '').strip()
    ident = str(ws.cell(row=r, column=3).value or '').strip()
    chave = (d.day, e if e else -s)
    planilha.append(chave)
    rotulo.setdefault(chave, []).append('L%d %s %s' % (r, ident, hist[:58]))

cs, cp = Counter(sistema), Counter(planilha)

so_sistema = list((cs - cp).elements())
so_planilha = list((cp - cs).elements())

print('=' * 96)
print('CONCILIACAO 34 (sistema) x PLANILHA %s' % ABA)
print('=' * 96)
print('linhas: sistema %d | planilha %d | casaram exatamente %d'
      % (len(sistema), len(planilha), len(list((cs & cp).elements()))))
print()


def fmt(v):
    return ('{:+,.2f}'.format(v / 100)).replace(',', 'X').replace('.', ',').replace('X', '.')


# --- procura par deslocado no tempo ---
print('-' * 96)
print('SO NO SISTEMA (%d linhas) — o sistema tem, a planilha nao' % len(so_sistema))
print('-' * 96)
restantes_p = Counter(so_planilha)
casados_data = []
sobra_sistema = []
for dia, val in sorted(so_sistema, key=lambda x: (x[0], -abs(x[1]))):
    achou = None
    for d2 in sorted(set(k[0] for k in restantes_p if k[1] == val), key=lambda d: abs(d - dia)):
        if abs(d2 - dia) <= 5:
            achou = d2
            break
    if achou is not None:
        restantes_p[(achou, val)] -= 1
        if restantes_p[(achou, val)] == 0:
            del restantes_p[(achou, val)]
        casados_data.append((dia, achou, val))
    else:
        sobra_sistema.append((dia, val))
        print('  %02d/08  %14s   <sem par na planilha>' % (dia, fmt(val)))
if not sobra_sistema:
    print('  (nenhuma)')

print()
print('-' * 96)
print('MESMO VALOR, DATA DIFERENTE (%d linhas) — nao e dinheiro faltando' % len(casados_data))
print('-' * 96)
for dia_s, dia_p, val in sorted(casados_data, key=lambda x: (x[0], -abs(x[2]))):
    etiq = rotulo.get((dia_p, val), [''])[0]
    print('  sistema %02d/08 | planilha %02d/08 | %14s | %s' % (dia_s, dia_p, fmt(val), etiq))

print()
print('-' * 96)
print('SO NA PLANILHA (%d linhas) — a planilha tem, o sistema nao' % sum(restantes_p.values()))
print('-' * 96)
for (dia, val), q in sorted(restantes_p.items(), key=lambda kv: (kv[0][0], -abs(kv[0][1]))):
    for etiq in rotulo.get((dia, val), [''])[:q]:
        print('  %02d/08  %14s   %s' % (dia, fmt(val), etiq))

print()
print('-' * 96)
print('EFEITO NO SALDO')
print('-' * 96)
ss = sum(v for _, v in sobra_sistema)
sp = sum(v * q for (_, v), q in restantes_p.items())
print('  so no sistema  %14s' % fmt(ss))
print('  so na planilha %14s' % fmt(sp))
print('  diferenca      %14s' % fmt(ss - sp))
