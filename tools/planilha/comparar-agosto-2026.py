# -*- coding: utf-8 -*-
"""Compara a aba Agosto-2026 da planilha do Itaú (Acop Files) com a conciliação
34 do sistema, dia a dia.

Leitura pura dos dois lados. O sistema é passado num TSV colado do navegador.
"""
import io
import sys
from collections import defaultdict

import openpyxl

CAMINHO = r'K:\GERAL\TI\Conciliação Itaú Acop Files.xlsx'
ABA = 'Agosto-2026'

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(CAMINHO, data_only=True)
ws = wb[ABA]


def cents(v):
    if v is None:
        return 0
    if isinstance(v, str):
        v = v.strip().replace('.', '').replace(',', '.')
        if not v:
            return 0
        return int(round(float(v) * 100))
    return int(round(float(v) * 100))


dias = defaultdict(lambda: {'n': 0, 'e': 0, 's': 0})
linhas = []
sem_data = []

for r in range(5, ws.max_row + 1):
    data = ws.cell(row=r, column=1).value
    doc = ws.cell(row=r, column=2).value
    ident = ws.cell(row=r, column=3).value
    hist = ws.cell(row=r, column=4).value
    ent = cents(ws.cell(row=r, column=5).value)
    sai = cents(ws.cell(row=r, column=6).value)

    if ent == 0 and sai == 0:
        continue

    reg = {
        'linha': r,
        'doc': str(doc).strip() if doc else '',
        'id': str(ident).strip() if ident else '',
        'hist': str(hist).strip() if hist else '',
        'e': ent,
        's': sai,
    }

    if data is None:
        sem_data.append(reg)
        continue

    d = data.strftime('%d/%m/%Y')
    reg['data'] = d
    linhas.append(reg)
    dias[d]['n'] += 1
    dias[d]['e'] += ent
    dias[d]['s'] += sai

abertura = cents(ws.cell(row=4, column=7).value)

print('PLANILHA %s' % ABA)
print('  saldo inicial      %15.2f' % (abertura / 100))
print('  linhas com data    %d' % len(linhas))
print('  linhas SEM data    %d  (bloco "em aberto")' % len(sem_data))
print('  total entradas     %15.2f' % (sum(l['e'] for l in linhas) / 100))
print('  total saidas       %15.2f' % (sum(l['s'] for l in linhas) / 100))

fecha = abertura + sum(l['e'] for l in linhas) - sum(l['s'] for l in linhas)
print('  fechamento apurado %15.2f' % (fecha / 100))
print()
print('POR DIA')
for d in sorted(dias, key=lambda x: (x[6:], x[3:5], x[:2])):
    v = dias[d]
    print('  %s n=%-3d E=%12.2f S=%12.2f' % (d, v['n'], v['e'] / 100, v['s'] / 100))

print()
print('BLOCO EM ABERTO (sem data) — %d linhas' % len(sem_data))
tot_e = sum(l['e'] for l in sem_data)
tot_s = sum(l['s'] for l in sem_data)
print('  entradas %12.2f   saidas %12.2f' % (tot_e / 100, tot_s / 100))
for l in sem_data:
    print('  L%-4d %-10s %-12s %-70s E=%10.2f S=%10.2f' % (
        l['linha'], l['doc'], l['id'], l['hist'][:70], l['e'] / 100, l['s'] / 100))
