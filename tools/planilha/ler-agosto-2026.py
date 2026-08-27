# -*- coding: utf-8 -*-
"""Despeja a aba Agosto-2026 da planilha de conciliação do Itaú da Acop Files.

Leitura pura: não escreve nada, não altera a planilha. Serve para conferir a
conciliação do sistema contra o que a equipe financeira mantém à mão.
"""
import io
import sys

import openpyxl

CAMINHO = r'K:\GERAL\TI\Conciliação Itaú Acop Files.xlsx'
ABA = 'Agosto-2026'

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(CAMINHO, data_only=True)
ws = wb[ABA]

print('DIM %s  max_row=%d max_col=%d' % (ws.dimensions, ws.max_row, ws.max_column))
print('-' * 100)

for linha in range(1, ws.max_row + 1):
    valores = []
    for col in range(1, 8):
        v = ws.cell(row=linha, column=col).value
        if v is None:
            valores.append('')
        elif hasattr(v, 'strftime'):
            valores.append(v.strftime('%d/%m/%Y'))
        elif isinstance(v, float):
            valores.append(('%.2f' % v))
        else:
            valores.append(str(v).strip())
    if any(valores):
        print('%3d | %s' % (linha, ' | '.join(valores)))
