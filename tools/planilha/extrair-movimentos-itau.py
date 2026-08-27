# -*- coding: utf-8 -*-
"""
Extrai de uma aba das planilhas de conciliação do Itaú as linhas que NÃO existem
no sistema, no formato que `gestao:importar-movimentos-planilha` consome.

POR QUE ESTE ARQUIVO ESTÁ NO REPOSITÓRIO
    A primeira extração (2026-08-24) foi feita por um script de rascunho que se
    perdeu junto com a sessão. O `import_key` daqueles 448 movimentos não é
    reconstituível, e sem o extrator não havia como reimportar nada sem duplicar.
    Um script que decide o que entra num sistema financeiro é parte do sistema.

O CRITÉRIO SÃO DUAS TRAVAS, E NENHUMA DAS DUAS BASTA SOZINHA
    (1) Linha COM id de origem na coluna C nunca é candidata. Ela nasceu em
        `contas`/`contasareceber` e chega pela sincronização.

        Sozinha, esta trava não basta: a leitura tentadora "ID vazio = digitado à
        mão = importa" está errada e custa caro. Em 2026 há 1.115 linhas sem ID e
        197 delas (17,7%) são liquidações que o sistema já tem — Sírio Libanês,
        Unicamp, FCM. Importar por ID vazio duplicaria centenas de milhares.

    (2) Linha SEM id só é candidata quando o sistema não tem nada de mesmo sentido
        e mesmo valor numa janela de ±7 dias — conferindo contra as liquidações E
        contra os movimentos manuais já importados. É o que torna o script seguro
        de rodar de novo: o que já entrou aparece no export e deixa de ser
        candidato sozinho.

        Sozinha, esta trava também não basta, e o caso que prova isso é a FOLHA.
        A planilha rateia o que a origem consolida: uma baixa de R$ 15.587,48 vira
        sete linhas por funcionário (Cartão, Poupança, um por banco). Nenhuma das
        sete acha par de mesmo valor, então a janela de ±7 dias as declararia
        ausentes e o import pagaria a folha duas vezes. O que as salva é a trava
        (1): todas carregam o id da baixa.

        Quando o rateio não fecha com a baixa (agosto/2026: faltam R$ 2.933,48 na
        folha e R$ 1.305,88 no adiantamento) isso é divergência a apurar com quem
        preenche a planilha — nunca dinheiro a importar.

USO
    1. exporte o sistema (somente leitura, no servidor):
         .codex-tmp/remote/exportar-acop-agosto.ps1
    2. python extrair-movimentos-itau.py <planilha.xlsx> <aba> <export.json> \
           --empresa "Acop Files" --saida candidatos.json
    3. confira SEMPRE em modo seco antes de gravar:
         php artisan gestao:importar-movimentos-planilha candidatos.json --dry-run
"""
import argparse
import datetime
import hashlib
import json
import sys

import openpyxl

# Tolerância entre a data do banco e a data de pagamento da origem. Vem do
# levantamento de 2026: das 1.072 linhas sem casamento exato no ano, 699 tinham
# par idêntico a até 3 dias e 66 entre 4 e 7 — depois disso a cauda some.
JANELA_DIAS = 7


def centavos(valor):
    return None if valor is None else int(round(float(valor) * 100))


def import_key(empresa, aba, linha, data, direcao, cents):
    """
    Identidade da linha: conteúdo + posição na aba.

    A posição entra porque duas tarifas idênticas no mesmo dia são dois fatos
    diferentes, e o conteúdo entra porque corrigir um valor na planilha deve
    gerar uma linha nova, não silenciosamente reaproveitar a antiga.
    """
    cru = f"v2|{empresa}|{aba}|{linha}|{data}|{direcao}|{cents}"
    return "xlsx:" + hashlib.sha1(cru.encode("utf-8")).hexdigest()


def ler_aba(caminho, aba):
    """Linhas de movimento da aba: as que têm data e valor. O bloco final sem
    data são títulos que ainda não caíram no banco — não são movimento."""
    ws = openpyxl.load_workbook(caminho, data_only=True)[aba]
    linhas = []

    for r in range(5, ws.max_row + 1):
        data, doc, idc, hist, ent, sai = (ws.cell(r, i).value for i in range(1, 7))

        if not isinstance(data, datetime.datetime):
            continue
        if ent is None and sai is None:
            continue

        linhas.append({
            "linha": r,
            "data": data.strftime("%Y-%m-%d"),
            "documento": "" if doc is None else str(doc).strip(),
            "id_origem": "" if idc is None else str(idc).strip(),
            "historico": "" if hist is None else str(hist).strip(),
            "direcao": "IN" if ent is not None else "OUT",
            "cents": centavos(ent if ent is not None else sai),
        })

    return linhas


def ler_sistema(caminho):
    """Liquidações e movimentos manuais que o sistema já tem, achatados no mesmo
    formato (data, sentido, centavos) para poderem ser confrontados."""
    dados = json.load(open(caminho, encoding="utf-8"))
    fatos = []

    for liq in dados.get("liquidacoes", []):
        fatos.append({
            "data": liq["settlement_date"][:10],
            "direcao": "OUT" if liq["type"] == "PAYMENT" else "IN",
            "cents": centavos(liq["amount"]),
        })

    for mov in dados.get("movimentos", []):
        fatos.append({
            "data": mov["movement_date"][:10],
            "direcao": mov["direction"],
            "cents": centavos(mov["amount"]),
        })

    return fatos


def candidatas(linhas, fatos, forcadas=frozenset()):
    """Separa em (candidatas, já cobertas, vindas da origem). Cada fato do sistema
    é consumido uma única vez: cinco linhas de R$ 1.166,00 no mesmo dia precisam
    de cinco fatos, não de um que sirva para todas."""
    usados = set()
    novas, cobertas, da_origem = [], [], []

    for linha in linhas:
        # Trava (1): tem id de origem, veio pela sincronização. Sai antes de
        # qualquer comparação de valor — é o que impede a folha rateada de ser
        # confundida com dinheiro ausente.
        #
        # `forcadas` é a saída deliberada para o caso em que a origem consolida o
        # que o banco pagou em várias contas: a baixa da folha é uma só, mas o
        # extrato do Itaú mostra apenas a parte que saiu DALI. Nesse caso a
        # planilha é o fato bancário e a baixa da origem tem de ser retirada da
        # conciliação junto — importar sem retirar conta o dinheiro duas vezes.
        if linha["id_origem"] and linha["linha"] not in forcadas:
            da_origem.append(linha)
            continue

        dia = datetime.date.fromisoformat(linha["data"])
        par = None

        for tol in range(JANELA_DIAS + 1):
            for i, fato in enumerate(fatos):
                if i in usados:
                    continue
                if fato["direcao"] != linha["direcao"] or fato["cents"] != linha["cents"]:
                    continue
                if abs((datetime.date.fromisoformat(fato["data"]) - dia).days) <= tol:
                    par = i
                    break
            if par is not None:
                break

        if par is None:
            novas.append(linha)
        else:
            usados.add(par)
            cobertas.append(linha)

    return novas, cobertas, da_origem


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("planilha")
    p.add_argument("aba")
    p.add_argument("export_sistema", help="JSON gerado pelo export somente-leitura do servidor")
    p.add_argument("--empresa", required=True, help="Nome exato como está em `contas`")
    p.add_argument("--saida", required=True)
    p.add_argument(
        "--forcar-linhas",
        default="",
        help="Números de linha da aba, separados por vírgula, que entram MESMO tendo id de "
             "origem. Só para o caso da origem consolidar o que o banco separou (folha). "
             "Exige retirar a baixa correspondente da conciliação, senão conta em dobro.",
    )
    args = p.parse_args()

    forcadas = frozenset(
        int(n) for n in args.forcar_linhas.replace(" ", "").split(",") if n
    )

    linhas = ler_aba(args.planilha, args.aba)
    fatos = ler_sistema(args.export_sistema)
    novas, cobertas, da_origem = candidatas(linhas, fatos, forcadas)

    registros = [{
        "empresa": args.empresa,
        "aba": args.aba,
        "linha_planilha": linha["linha"],
        "movement_date": linha["data"],
        "direction": linha["direcao"],
        "amount": f"{linha['cents'] / 100:.2f}",
        "document_number": linha["documento"] or None,
        "history": linha["historico"],
        "import_key": import_key(
            args.empresa, args.aba, linha["linha"], linha["data"], linha["direcao"], linha["cents"],
        ),
    } for linha in novas]

    json.dump(registros, open(args.saida, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    print(f"aba {args.aba}: {len(linhas)} linhas de movimento")
    print(f"  vindas da origem (tem id): {len(da_origem)}")
    print(f"  já cobertas pelo sistema : {len(cobertas)}")
    print(f"  candidatas a importar    : {len(novas)}  ->  {args.saida}")

    for linha in novas:
        sinal = "+" if linha["direcao"] == "IN" else "-"
        print(f"    l{linha['linha']:<4} {linha['data']}  {sinal}{linha['cents'] / 100:>12,.2f}  {linha['historico'][:58]}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
